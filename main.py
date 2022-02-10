from openpyxl import Workbook, load_workbook, drawing
import datetime
import PIL.Image
import locale
import holidays


class TimesheetGenerator:
    def __init__(self):
        self.workbook = load_workbook("template.xlsx")
        self.worksheet = self.workbook.active

    def run(self):
        self.day_col_offset = 9
        self.locale = 'de_DE.UTF-8'
        self.holidays = holidays.Germany(prov="BY")
        self.holiday_note = "Feiertag"
        self.weekend_days = [5, 6]
        self.fill()

        self.sign_image_path = "sign.png"
        self.sign_image_width = 205
        self.sign_image_pos = "E37"
        self.sign_note_row = 36
        self.sign_note_column = 5
        self.sign_note_city = "Musterstadt"
        self.sign()

        self.worksheet.protection.enable()

    def fill(self):
        locale.setlocale(locale.LC_ALL, self.locale)

        first_day = datetime.date.today().replace(day=1)
        last_day = self.get_last_day_in_month(first_day)

        day = first_day
        days = []
        while day <= last_day:
            if day.weekday() in self.weekend_days:
                day += datetime.timedelta(days=1)
                continue
            if day in self.holidays:
                days.append((day, self.holidays.get(day)))
                day += datetime.timedelta(days=1)
                continue
            days.append((day, ""))
            day += datetime.timedelta(days=1)

        for row, (day, comment) in enumerate(days):
            self.worksheet.cell(row=row+self.day_col_offset,
                                column=1,
                                value=day)
            if comment != "":
                self.worksheet.cell(row=row+self.day_col_offset,
                                    column=2, value=self.holiday_note)
                self.worksheet.cell(row=row+self.day_col_offset,
                                    column=3, value=f"")
                self.worksheet.cell(row=row+self.day_col_offset,
                                    column=4, value=f"")
                self.worksheet.cell(row=row+self.day_col_offset,
                                    column=5, value=f"")
            else:
                self.worksheet.cell(
                    row=row+self.day_col_offset, column=2, value="")
                self.worksheet.cell(row=row+self.day_col_offset, column=3,
                                    value=datetime.time(10))
                self.worksheet.cell(row=row+self.day_col_offset, column=4,
                                    value=datetime.time(14, 30))
                self.worksheet.cell(row=row+self.day_col_offset, column=5,
                                    value=datetime.time(0, 30))

            day += datetime.timedelta(days=1)

    def get_last_day_in_month(self, day):
        next_month = day.replace(day=28) + datetime.timedelta(days=4)
        return next_month - datetime.timedelta(days=next_month.day)

    def get_last_work_day_in_month(self, day):
        day = self.get_last_day_in_month(day)
        while day.weekday() in self.weekend_days or day in self.holidays:
            day -= datetime.timedelta(days=1)
        return day

    def sign(self):
        sign_day = self.get_last_work_day_in_month(datetime.date.today())
        sign_note = f"{self.sign_note_city}, {sign_day.strftime('%a, %d.%m.%Y')}"
        self.worksheet.cell(row=self.sign_note_row,
                            column=self.sign_note_column,
                            value=sign_note)

        fp = open(self.sign_image_path, "rb")
        img = PIL.Image.open(fp)

        wpercent = (self.sign_image_width/float(img.size[0]))
        height = int(float(img.size[1])*float(wpercent))

        resizedImg = img.resize(
            (self.sign_image_width, height), PIL.Image.ANTIALIAS)
        fp.seek(0)
        resizedImg.fp = fp

        table_img = drawing.image.Image(resizedImg)
        table_img.anchor = self.sign_image_pos
        self.worksheet.add_image(table_img)

    def save(self, path):
        self.workbook.save(path)


def main():
    gen = TimesheetGenerator()
    gen.run()
    gen.save("test_out.xlsx")


if __name__ == "__main__":
    main()
